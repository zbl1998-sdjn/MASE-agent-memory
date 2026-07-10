"""CSV/Excel 表格摄入:确定性抽取契约 + 端到端溯源(零 LLM,全可复现)。

契约(与 kv_extract/structure_facts 同族的确定性抽取):
- 两列表 = KV 形态:每行一条 fact(key=第一列,value=第二列);
- ≥3 列表 = 行打包:key=首格,value=该行完整渲染切片(structure_facts
  表格行打包同族);
- 值逐字出现在 full_text → 治理层 evidence 机械 span 定位天然通过;
- 空表/全空行零产出(halluc_ok 模式);每 sheet 事实数上限护栏,超出
  记 warning 不静默截断。
"""
from __future__ import annotations

import sys
from pathlib import Path

import pytest

_ROOT = Path(__file__).resolve().parents[1]
for _p in (_ROOT / "src", _ROOT):
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))

from mase.multimodal.document_loader import load_media
from mase.multimodal.security import classify_media
from mase.multimodal.tabular_extractor import TabularExtractor


def _extract(tmp_path: Path, filename: str, content: bytes):
    """写入文件 → 分类 → 载荷 → 抽取,返回 ExtractionResult。"""
    from mase.multimodal.extractor import MediaAssetInfo

    path = tmp_path / filename
    path.write_bytes(content)
    media_type = classify_media(path)
    payload = load_media(path, media_type)
    asset = MediaAssetInfo(
        media_id=1, sha256="deadbeef" * 8, media_type=media_type,
        source_uri=filename, page_count=0,
    )
    return TabularExtractor().extract(asset, payload)


class TestSecurityAllowlist:
    def test_csv_and_xlsx_are_classified(self, tmp_path):
        csv_path = tmp_path / "a.csv"
        csv_path.write_bytes(b"k,v\n")
        assert classify_media(csv_path) == "text/csv"

    def test_xlsx_mime(self, tmp_path):
        import openpyxl

        path = tmp_path / "b.xlsx"
        wb = openpyxl.Workbook()
        wb.save(path)
        assert classify_media(path) == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


class TestCsvExtraction:
    def test_two_column_csv_yields_kv_facts_with_verbatim_values(self, tmp_path):
        content = "报销上限,500元\n项目代号,Phoenix\n".encode()
        result = _extract(tmp_path, "kv.csv", content)
        pairs = {f.key: f.value for f in result.candidate_facts}
        assert pairs["报销上限"] == "500元"
        assert pairs["项目代号"] == "Phoenix"
        for fact in result.candidate_facts:
            assert fact.value in result.full_text  # 治理 span 定位前提
            assert fact.evidence in result.full_text

    def test_multi_column_csv_packs_rows(self, tmp_path):
        content = (
            "姓名,部门,工号\n"
            "孙艺,财务部,200652468\n"
            "李明,工程部,200701001\n"
        ).encode()
        result = _extract(tmp_path, "roster.csv", content)
        by_key = {f.key: f for f in result.candidate_facts}
        assert "孙艺" in by_key
        assert "财务部" in by_key["孙艺"].value
        assert "200652468" in by_key["孙艺"].value
        for fact in result.candidate_facts:
            assert fact.value in result.full_text

    def test_gbk_encoded_csv_is_decoded(self, tmp_path):
        content = "供应商,山东蓝天贸易\n".encode("gbk")
        result = _extract(tmp_path, "gbk.csv", content)
        pairs = {f.key: f.value for f in result.candidate_facts}
        assert pairs["供应商"] == "山东蓝天贸易"
        assert result.metadata is not None
        assert result.metadata.get("encoding") == "gbk"

    def test_empty_csv_yields_zero_facts(self, tmp_path):
        result = _extract(tmp_path, "empty.csv", b"\n\n")
        assert result.candidate_facts == ()
        assert result.warnings == ()

    def test_fact_cap_warns_instead_of_silent_truncation(self, tmp_path):
        rows = "\n".join(f"key{i},value{i}" for i in range(300))
        result = _extract(tmp_path, "big.csv", rows.encode("utf-8"))
        assert len(result.candidate_facts) == 200
        assert any("300" in w and "200" in w for w in result.warnings)


class TestXlsxExtraction:
    def test_multi_sheet_xlsx_extracts_all_sheets(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "预算"
        ws1.append(["季度预算", "120000"])
        ws2 = wb.create_sheet("联系人")
        ws2.append(["姓名", "电话", "城市"])
        ws2.append(["王芳", "13800001111", "济南"])
        path = tmp_path / "book.xlsx"
        wb.save(path)

        result = _extract(tmp_path, "book.xlsx", path.read_bytes())
        pairs = {f.key: f for f in result.candidate_facts}
        assert pairs["季度预算"].value == "120000"
        assert "13800001111" in pairs["王芳"].value
        assert "预算" in result.full_text and "联系人" in result.full_text
        for fact in result.candidate_facts:
            assert fact.value in result.full_text

    def test_numeric_cells_render_deterministically(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["总额", 1234.5])
        ws.append(["数量", 42])
        path = tmp_path / "nums.xlsx"
        wb.save(path)

        result = _extract(tmp_path, "nums.xlsx", path.read_bytes())
        pairs = {f.key: f.value for f in result.candidate_facts}
        assert pairs["总额"] == "1234.5"
        assert pairs["数量"] == "42"  # int 不得渲染成 42.0


class TestIngestEndToEnd:
    def test_ingest_folder_routes_csv_through_tabular_extractor(self, tmp_path, monkeypatch):
        monkeypatch.delenv("MASE_MEMORY_DIR", raising=False)
        monkeypatch.setenv("MASE_DB_PATH", str(tmp_path / "ingest.db"))
        docs = tmp_path / "docs"
        docs.mkdir()
        (docs / "budget.csv").write_text("季度预算,120000\n", encoding="utf-8")

        from mase.multimodal.ingest import ingest_folder

        report = ingest_folder(docs, asset_root=tmp_path / "assets")
        assert report.processed == ("budget.csv",)
        assert report.infra_errors == ()
        assert report.facts_written == 1
        # 治理层双写:值逐字在 full_text 中 → span 定位必须成功,零 warning。
        assert report.facts_governed == 1
        assert report.governance_warnings == ()

    def test_ingest_is_idempotent_for_tabular(self, tmp_path, monkeypatch):
        monkeypatch.delenv("MASE_MEMORY_DIR", raising=False)
        monkeypatch.setenv("MASE_DB_PATH", str(tmp_path / "ingest.db"))
        docs = tmp_path / "docs"
        docs.mkdir()
        (docs / "budget.csv").write_text("季度预算,120000\n", encoding="utf-8")

        from mase.multimodal.ingest import ingest_folder

        first = ingest_folder(docs, asset_root=tmp_path / "assets")
        second = ingest_folder(docs, asset_root=tmp_path / "assets")
        assert first.extractions == 1
        assert second.extractions == 0
        assert any(s["reason"] == "already_extracted" for s in second.skipped)


class TestMissingDependency:
    def test_xlsx_without_openpyxl_raises_actionable_error(self, tmp_path, monkeypatch):
        import builtins

        real_import = builtins.__import__

        def _no_openpyxl(name, *args, **kwargs):
            if name == "openpyxl":
                raise ImportError("No module named 'openpyxl'")
            return real_import(name, *args, **kwargs)

        monkeypatch.setattr(builtins, "__import__", _no_openpyxl)

        from mase.multimodal.document_loader import MissingDependencyError

        with pytest.raises(MissingDependencyError, match="openpyxl"):
            _extract(tmp_path, "x.xlsx", b"PK\x03\x04fake")
